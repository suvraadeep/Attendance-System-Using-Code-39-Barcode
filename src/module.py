import win32com.client
import cv2
import time
from pyzbar import pyzbar
from tkinter import messagebox
import pythoncom
from datetime import datetime
import openpyxl as xl
from openpyxl import load_workbook


def protected_view_excel(path,sheets):  #opensheetsofexcelinprotectedview
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    xl.DisplayAlerts = False
    workbook = xl.Workbooks.Open(path)
    active_sheet = workbook.Worksheets(sheets)
    active_sheet.Activate()
    for sheet in workbook.Worksheets:
        if sheet.Name != sheets:
            sheet.Visible = 0
    workbook.Unprotect(Password="password")
    workbook.Protect(Password="password", Structure=True, Windows=True, Readonly=True)
    xl.ActiveWindow.Activate()
    workbook.Close(False)
    xl.Quit()
    pythoncom.CoUninitialize()

def write_date_to_sub(path, sheetname, date):
    book = load_workbook(path)
    sheet = book[sheetname]
    last_date_cell = None
    for col in range(4, 30):
        if sheet.cell(row=1, column=col).value:
            last_date_cell = sheet.cell(row=1, column=col)
            if last_date_cell.value == date.strftime("%d-%m-%Y"):
                return  
    if last_date_cell:
        next_date_cell = last_date_cell.offset(column=1)
        if not next_date_cell.value:
            next_date_cell.value = date.strftime("%d-%m-%Y")
            sheet.parent.save(path)
            return
    for col in range(4, 30):
        if not sheet.cell(row=1, column=col).value:
            sheet.cell(row=1, column=col).value = date.strftime("%d-%m-%Y")
            sheet.parent.save(path)
            return

def scan_barcodes(timeout):         #scanningbarcode
    cap = cv2.VideoCapture(0)
    barcodes = []
    start_time = time.time()
    while True:
        ret, frame = cap.read()
        decoded_objects = pyzbar.decode(frame)
        for obj in decoded_objects:
            barcodes == obj.data
            (x, y, w, h) = obj.rect
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            if obj.data.decode("utf-8") not in barcodes:
                barcodes.append(obj.data.decode("utf-8"))
        cv2.imshow('Barcode Scanner [Click Space Bar to Exit the Camera]', frame)
        if time.time() - start_time > timeout:
            break
        if cv2.waitKey(1) & 0xFF == ord(" "):
            break
    cap.release()
    cv2.destroyAllWindows()
    return barcodes

def add_p(num,path,sheet):     #addingdate
    wb = xl.load_workbook(path)
    sheet = wb[sheet]
    today = datetime.today().strftime('%d-%m-%Y')
    date_col_index = None
    num_row_index = None
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == today:
            date_col_index = col
            break
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=2).value == num:
            num_row_index = row
            break
    if date_col_index and num_row_index:
        sheet.cell(row=num_row_index, column=date_col_index).value = "P"
        wb.save(path)
        messagebox.showinfo("SUCCESS", "Attendance added successfully")
    else:
        messagebox.showerror("ERROR", "You are not able to add attendance. \nTry Again!")

def absenttopresent(path,sheet,date,num):   #changesfromabsenttopresent
    wb = xl.load_workbook(path)
    sheet = wb[sheet]
    date_col_index = None
    num_row_index = None
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == date:
            date_col_index = col
            break
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=2).value == num:
            num_row_index = row
            break
    if date_col_index and num_row_index:
        sheet.cell(row=num_row_index, column=date_col_index).value = "P"
        wb.save(path)
        messagebox.showinfo("SUCCESS", "ABSENT modified to PRESENT \nSuccessfully")
    else:
        messagebox.showerror("ERROR","ERROR OCCURED \nTRY AGAIN")

def presenttoabsent(path,sheet,date,num):      #changesfrompresenttoabsent
    wb = xl.load_workbook(path)
    sheet = wb[sheet]
    date_col_index = None
    num_row_index = None
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == date:
            date_col_index = col
            break
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=2).value == num:
            num_row_index = row
            break
    if date_col_index and num_row_index:
        sheet.cell(row=num_row_index, column=date_col_index).value = "A"
        wb.save(path)
        messagebox.showinfo("SUCCESS", "PRESENT modified to ABSENT \nSuccessfully")
    else:
        messagebox.showerror("ERROR","ERROR OCCURED \nTRY AGAIN")

def markpresent(path,sheet,date,nums):
    wb = xl.load_workbook(path)
    sheet = wb[sheet]
    date_col_index = None
    num_row_indices = []
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == date.strftime("%d-%m-%Y"):
            date_col_index = col
            break
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=1).value % 100 in nums:
            num_row_indices.append(row)
    if date_col_index and num_row_indices:
        for num_row_index in num_row_indices:
            sheet.cell(row=num_row_index, column=date_col_index).value = "P"
        wb.save(path)
        messagebox.showinfo("SUCCESS", "Present marked successfully")
    else:
        messagebox.showerror("ERROR","ERROR OCCURED \nTRY AGAIN")

def markabsent(path,sheet,date,nums):
    wb = xl.load_workbook(path)
    sheet = wb[sheet]
    date_col_index = None
    num_row_indices = []
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == date.strftime("%d-%m-%Y"):
            date_col_index = col
            break
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=1).value % 100 in nums:
            num_row_indices.append(row)
    if date_col_index and num_row_indices:
        for num_row_index in num_row_indices:
            sheet.cell(row=num_row_index, column=date_col_index).value = "A"
        wb.save(path)
        messagebox.showinfo("SUCCESS", "Absent marked successfully")
    else:
        messagebox.showerror("ERROR","ERROR OCCURED \nTRY AGAIN")